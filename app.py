from __future__ import annotations

import calendar
import csv
import io
import re
import unicodedata
import zipfile
from typing import Callable, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from dateutil import parser as dtparser


# ---------- Utilities ----------

def _parse_money(val) -> float:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return 0.0
    if isinstance(val, (int, float, np.number)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1].strip()
    if s.endswith("-"):
        neg, s = True, s[:-1].strip()
    s = re.sub(r"(idr|rp|cr|dr)", "", s, flags=re.IGNORECASE)
    s = re.sub(r"[^0-9\.,\-]", "", s).strip()
    if s.startswith("-"):
        neg, s = True, s[1:].strip()

    last_dot = s.rfind(".")
    last_com = s.rfind(",")
    if last_dot == -1 and last_com == -1:
        num_s = s
    elif last_dot > last_com:
        num_s = s.replace(",", "")
    else:
        num_s = s.replace(".", "").replace(",", ".")

    try:
        num = float(num_s)
    except Exception:
        num_s = s.replace(".", "").replace(",", "")
        num = float(num_s) if num_s else 0.0

    return -num if neg else num


def _to_num(sr: pd.Series) -> pd.Series:
    return sr.apply(_parse_money).astype(float)


def _norm_str(val) -> str:
    s = "" if val is None else str(val)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().lower()



def _is_bca_va_product(prod_val) -> bool:
    """Return True for BCA Virtual Account products (incl. blu variants).

    Business rules:
    - Any product that clearly indicates BCA VA (contains 'va'/'virtual account' and 'bca') is BCA VA.
    - blu-branded BCA products that may not include the string 'VA' (e.g. 'blu BCA Online',
      'blu by BCA Digital') are also treated as BCA VA.
    """
    s = _norm_str(prod_val)

    # Exact/near-exact known labels found in settlement reports.
    if s in {
        "bca va online",
        "blu bca va online",
        "blu bca",
        "blu bca online",
        "blu by bca digital",
        "blu by bca",
    }:
        return True

    # General signals.
    has_bca = "bca" in s
    has_va = ("va" in s) or ("virtual account" in s)

    # blu + bca is treated as BCA VA even if 'VA' text is missing.
    has_blu_bca = ("blu" in s) and has_bca

    return has_bca and (has_va or has_blu_bca)


def _norm_bank(val) -> str:
    """Normalize bank names; treat 'blu BCA' as 'bca'."""
    s = _norm_str(val)
    if "bca" in s:
        return "bca"
    return s


def _norm_order_id(val) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return ""
    s = re.sub(r"\s+", "", s)
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s.casefold()


def _first_nonempty_text(sr: pd.Series) -> str:
    for v in sr:
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s and s.lower() not in {"nan", "none", "null"}:
            return s
    return ""


_ddmmyyyy = re.compile(r"\b(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{2,4})\b")


def _to_date(val) -> Optional[pd.Timestamp]:
    if pd.isna(val):
        return None

    if isinstance(val, (int, float, np.number)):
        if not np.isfinite(val):
            return None
        if 1 <= float(val) <= 100000:
            try:
                base = pd.Timestamp("1899-12-30")
                return (base + pd.to_timedelta(float(val), unit="D")).normalize()
            except Exception:
                pass

    if isinstance(val, (pd.Timestamp, np.datetime64)):
        try:
            return pd.to_datetime(val).normalize()
        except Exception:
            return None

    s = str(val).strip()
    if not s:
        return None

    if re.fullmatch(r"\d{1,5}(\.0+)?", s):
        try:
            serial = float(s)
            if 1 <= serial <= 100000:
                base = pd.Timestamp("1899-12-30")
                return (base + pd.to_timedelta(serial, unit="D")).normalize()
        except Exception:
            pass

    explicit_formats = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    )
    for fmt in explicit_formats:
        try:
            return pd.Timestamp(pd.to_datetime(s, format=fmt, errors="raise").date())
        except Exception:
            continue

    m = _ddmmyyyy.search(s)
    if m:
        d, M, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        try:
            return pd.Timestamp(year=int(y), month=int(M), day=int(d))
        except Exception:
            pass

    for dayfirst in (True, False):
        try:
            d = dtparser.parse(s, dayfirst=dayfirst, fuzzy=True)
            return pd.Timestamp(d.date())
        except Exception:
            continue

    return None


def _fill_action_from_created(df: pd.DataFrame, action_col: str, created_col: str) -> pd.DataFrame:
    """Isi kolom Action dari kolom Created jika Action kosong.

    Output Action dibuat dd/mm/yyyy supaya parsing tanggal konsisten.
    Robust untuk Created yang bisa berupa:
    - yyyy-mm-dd hh:mm:ss
    - dd/mm/yyyy hh:mm:ss
    """
    if df is None or df.empty or not action_col or not created_col:
        return df

    action_raw = df[action_col]
    action_str = action_raw.astype(str).str.strip().str.lower()
    empty_mask = action_raw.isna() | action_str.eq("") | action_str.eq("nan") | action_str.eq("none")
    if not empty_mask.any():
        return df

    created_str = df[created_col].astype(str).str.strip()

    dt_dayfirst = pd.to_datetime(created_str, errors="coerce", dayfirst=True)
    dt_monthfirst = pd.to_datetime(created_str, errors="coerce", dayfirst=False)
    created_dt = dt_dayfirst.fillna(dt_monthfirst)

    filled = created_dt.dt.strftime("%d/%m/%Y")

    fallback = created_str.str.slice(0, 10)
    ymd = fallback.str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)
    fallback_norm = fallback.copy()
    fallback_norm.loc[ymd] = (
        fallback.loc[ymd].str.slice(8, 10) + "/"
        + fallback.loc[ymd].str.slice(5, 7) + "/"
        + fallback.loc[ymd].str.slice(0, 4)
    )

    filled = filled.where(created_dt.notna(), fallback_norm)

    df.loc[empty_mask, action_col] = filled.loc[empty_mask]
    return df


def _read_any(uploaded_file) -> pd.DataFrame:
    if not uploaded_file:
        return pd.DataFrame()
    name = uploaded_file.name.lower()

    try:
        if name.endswith(".csv"):
            for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1"):
                try:
                    uploaded_file.seek(0)
                    return pd.read_csv(
                        uploaded_file,
                        encoding=enc,
                        sep=None,
                        engine="python",
                        dtype=str,
                        na_filter=False,
                    )
                except Exception:
                    continue
            st.error(f"CSV gagal dibaca: {uploaded_file.name}. Simpan ulang sebagai UTF-8.")
            return pd.DataFrame()

        if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            uploaded_file.seek(0)
            return pd.read_excel(uploaded_file, engine="openpyxl")

        if name.endswith(".xls"):
            try:
                uploaded_file.seek(0)
                return pd.read_excel(uploaded_file, engine="xlrd")
            except Exception:
                pass
            try:
                uploaded_file.seek(0)
                raw = uploaded_file.read()
                from pyexcel_xls import get_data  # type: ignore

                book = get_data(io.BytesIO(raw))
                for _sh, rows in book.items():
                    if not rows:
                        continue
                    header = [str(x).strip() if x is not None else "" for x in rows[0]]
                    body = rows[1:] if len(rows) > 1 else []
                    return pd.DataFrame(body, columns=header)
            except Exception:
                pass
            try:
                uploaded_file.seek(0)
                return pd.read_excel(uploaded_file, engine="openpyxl")
            except Exception:
                st.error("Gagal membaca file .xls. Pasang 'xlrd' atau 'pyexcel-xls', atau simpan ulang ke .xlsx.")
                return pd.DataFrame()

        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file)

    except ImportError:
        st.error("Dukungan .xls perlu paket 'xlrd' atau 'pyexcel-xls'. Tambahkan di requirements.txt.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Gagal membaca {uploaded_file.name}: {e}")
        return pd.DataFrame()


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df.columns = [str(c).strip().lstrip("\ufeff") for c in df.columns]
    return df


def _read_csv_with_sep(uploaded_file, enc: str, sep, engine: Optional[str] = "python") -> pd.DataFrame:
    uploaded_file.seek(0)
    kwargs = {
        "encoding": enc,
        "sep": sep,
        "dtype": str,
        "na_filter": False,
    }
    if engine is not None:
        kwargs["engine"] = engine
    return pd.read_csv(uploaded_file, **kwargs)


def _read_ticket_csv_comma(uploaded_file, enc: str) -> pd.DataFrame:
    uploaded_file.seek(0)
    return pd.read_csv(
        uploaded_file,
        encoding=enc,
        sep=",",
        dtype=str,
        na_filter=False,
        low_memory=False,
    )



def _decode_uploaded_text(uploaded_file, enc: str) -> str:
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    if isinstance(raw, str):
        return raw
    return raw.decode(enc, errors="replace")


def _rows_to_dataframe(rows: List[List[str]]) -> pd.DataFrame:
    clean_rows: List[List[str]] = []
    for row in rows:
        vals = ["" if v is None else str(v) for v in row]
        if any(v.strip() for v in vals):
            clean_rows.append(vals)

    if not clean_rows:
        return pd.DataFrame()

    max_cols = max(len(r) for r in clean_rows)
    padded = [r + [""] * (max_cols - len(r)) for r in clean_rows]
    header = padded[0]
    body = padded[1:] if len(padded) > 1 else []
    return pd.DataFrame(body, columns=header)


def _ticket_header_markers() -> Tuple[str, ...]:
    return (
        "created",
        "action date",
        "action",
        "tarif",
        "bank",
        "st bayar",
        "status bayar",
        "order id",
        "type",
        "payment type",
        "payment channel",
        "channel",
    )


def _ticket_header_score(values) -> int:
    markers = _ticket_header_markers()
    norm_vals = [_norm_str(v) for v in values if str(v).strip()]
    score = 0
    for marker in markers:
        if any(marker in v for v in norm_vals):
            score += 1
    return score


def _promote_ticket_header(df: pd.DataFrame, scan_max: int = 20) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    current_score = _ticket_header_score(df.columns)
    if current_score >= 4:
        return _clean_columns(df)

    best_row = None
    best_score = current_score

    for r in range(min(scan_max, len(df))):
        row_vals = df.iloc[r].tolist()
        score = _ticket_header_score(row_vals)
        if score > best_score:
            best_score = score
            best_row = r

    if best_row is None or best_score < 4:
        return _clean_columns(df)

    cols = [str(x).strip() for x in df.iloc[best_row].tolist()]
    out = df.iloc[best_row + 1 :].copy()
    out.columns = cols
    return _clean_columns(out)


def _ticket_required_col_count(df: pd.DataFrame) -> int:
    if df is None or df.empty:
        return 0

    required_groups = [
        ["Created", "Created Date", "Created At", "Created Time"],
        ["Tarif", "tarif"],
        ["Bank", "Payment Channel", "channel", "payment method"],
        ["St Bayar", "Status Bayar", "status", "status bayar"],
    ]

    found = 0
    for names in required_groups:
        if _find_col(df, names) is not None:
            found += 1
    return found


def _ticket_df_quality_score(df: pd.DataFrame) -> Tuple[int, int, int, int]:
    if df is None or df.empty:
        return (-1, -1, -1, -1)

    required_count = _ticket_required_col_count(df)
    header_score = _ticket_header_score(df.columns)
    width = df.shape[1]

    sampled_cells = []
    if width > 0:
        sample_rows = min(20, len(df))
        for r in range(sample_rows):
            for v in df.iloc[r].tolist():
                s = str(v).strip()
                if s:
                    sampled_cells.append(s)

    data_hint_score = 0
    hint_markers = ("paid", "espay", "go show", "online", "bca", "bri", "mandiri")
    for marker in hint_markers:
        if any(marker in _norm_str(v) for v in sampled_cells):
            data_hint_score += 1

    return (required_count, header_score, width, data_hint_score)


def _ticket_csv_parse_is_valid(df: pd.DataFrame) -> bool:
    if df is None or df.empty:
        return False
    required_count, header_score, width, _ = _ticket_df_quality_score(df)
    if width <= 1:
        return False
    if required_count >= 4:
        return True
    return required_count >= 3 and header_score >= 4


def _ticket_csv_parse_is_suspicious(df: pd.DataFrame) -> bool:
    return not _ticket_csv_parse_is_valid(df)


def _read_ticket_csv_comma(uploaded_file, enc: str) -> pd.DataFrame:
    uploaded_file.seek(0)
    return pd.read_csv(
        uploaded_file,
        encoding=enc,
        sep=",",
        dtype=str,
        na_filter=False,
        low_memory=False,
    )


def _read_ticket_csv_comma_manual(uploaded_file, enc: str) -> pd.DataFrame:
    text_data = _decode_uploaded_text(uploaded_file, enc)

    candidates: List[pd.DataFrame] = []

    try:
        reader = csv.reader(io.StringIO(text_data), delimiter=",", quotechar='"', doublequote=True)
        rows = [row for row in reader]
        candidates.append(_rows_to_dataframe(rows))
    except Exception:
        pass

    raw_lines = text_data.splitlines()

    def _parse_lines(unwrap_outer_quotes: bool = False, simple_split: bool = False) -> pd.DataFrame:
        rows: List[List[str]] = []
        for line in raw_lines:
            if line is None:
                continue
            stripped = line.rstrip("\r\n")
            if not stripped.strip():
                continue

            working = stripped
            if unwrap_outer_quotes:
                s = working.strip()
                if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
                    s = s[1:-1]
                working = s

            try:
                if simple_split:
                    row = working.split(",")
                else:
                    row = next(csv.reader([working], delimiter=",", quotechar='"', doublequote=True))
            except Exception:
                row = working.split(",")

            rows.append(row)

        return _rows_to_dataframe(rows)

    candidates.append(_parse_lines(unwrap_outer_quotes=False, simple_split=False))
    candidates.append(_parse_lines(unwrap_outer_quotes=True, simple_split=False))
    candidates.append(_parse_lines(unwrap_outer_quotes=True, simple_split=True))

    best_df = pd.DataFrame()
    best_score = (-1, -1, -1)

    for candidate in candidates:
        candidate = _promote_ticket_header(candidate)
        score = _ticket_df_quality_score(candidate)
        if score > best_score:
            best_df = candidate
            best_score = score

    return _clean_columns(best_df)


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df.columns = [str(c).strip().lstrip("\ufeff") for c in df.columns]
    return df


def _read_csv_with_sep(uploaded_file, enc: str, sep, engine: Optional[str] = "python") -> pd.DataFrame:
    uploaded_file.seek(0)
    kwargs = {
        "encoding": enc,
        "sep": sep,
        "dtype": str,
        "na_filter": False,
    }
    if engine is not None:
        kwargs["engine"] = engine
    return pd.read_csv(uploaded_file, **kwargs)


def _read_tiket_detail_any(uploaded_file) -> pd.DataFrame:
    if not uploaded_file:
        return pd.DataFrame()

    name = uploaded_file.name.lower()

    try:
        if name.endswith(".csv"):
            best_df = pd.DataFrame()
            best_score = (-1, -1, -1, -1)

            for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1"):
                try:
                    pandas_df = _promote_ticket_header(_clean_columns(_read_ticket_csv_comma(uploaded_file, enc)))
                    pandas_score = _ticket_df_quality_score(pandas_df)

                    if _ticket_csv_parse_is_valid(pandas_df):
                        return pandas_df

                    if pandas_score > best_score:
                        best_df = pandas_df
                        best_score = pandas_score
                except Exception:
                    pass

                try:
                    manual_df = _promote_ticket_header(_clean_columns(_read_ticket_csv_comma_manual(uploaded_file, enc)))
                    manual_score = _ticket_df_quality_score(manual_df)

                    if _ticket_csv_parse_is_valid(manual_df):
                        return manual_df

                    if manual_score > best_score:
                        best_df = manual_df
                        best_score = manual_score
                except Exception:
                    pass

            if not best_df.empty:
                return best_df

            st.error(
                f"CSV Tiket Detail gagal dibaca: {uploaded_file.name}. "
                "Pastikan file CSV valid dengan delimiter koma (,)."
            )
            return pd.DataFrame()

        return _clean_columns(_read_any(uploaded_file))

    except Exception as e:
        st.error(f"Gagal membaca file Tiket Detail {uploaded_file.name}: {e}")
        return pd.DataFrame()


def _read_any(uploaded_file) -> pd.DataFrame:
    if not uploaded_file:
        return pd.DataFrame()
    name = uploaded_file.name.lower()

    try:
        if name.endswith(".csv"):
            for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1"):
                try:
                    return _clean_columns(_read_csv_with_sep(uploaded_file, enc, None))
                except Exception:
                    continue
            st.error(f"CSV gagal dibaca: {uploaded_file.name}. Simpan ulang sebagai UTF-8.")
            return pd.DataFrame()

        if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            uploaded_file.seek(0)
            return pd.read_excel(uploaded_file, engine="openpyxl")

        if name.endswith(".xls"):
            try:
                uploaded_file.seek(0)
                return pd.read_excel(uploaded_file, engine="xlrd")
            except Exception:
                pass
            try:
                uploaded_file.seek(0)
                raw = uploaded_file.read()
                from pyexcel_xls import get_data  # type: ignore

                book = get_data(io.BytesIO(raw))
                for _sh, rows in book.items():
                    if not rows:
                        continue
                    header = [str(x).strip() if x is not None else "" for x in rows[0]]
                    body = rows[1:] if len(rows) > 1 else []
                    return pd.DataFrame(body, columns=header)
            except Exception:
                pass
            try:
                uploaded_file.seek(0)
                return pd.read_excel(uploaded_file, engine="openpyxl")
            except Exception:
                st.error("Gagal membaca file .xls. Pasang 'xlrd' atau 'pyexcel-xls', atau simpan ulang ke .xlsx.")
                return pd.DataFrame()

        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file)

    except ImportError:
        st.error("Dukungan .xls perlu paket 'xlrd' atau 'pyexcel-xls'. Tambahkan di requirements.txt.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Gagal membaca {uploaded_file.name}: {e}")
        return pd.DataFrame()


def _find_col(df: pd.DataFrame, names: List[str]) -> Optional[str]:
    if df.empty:
        return None
    cols = [c for c in df.columns if isinstance(c, str)]
    m = {c.lower().strip().lstrip("\ufeff"): c for c in cols}

    for n in names:
        key = n.lower().strip()
        if key in m:
            return m[key]

    for n in names:
        key = n.lower().strip()
        for c in cols:
            if key in c.lower():
                return c

    return None


def _adjust_created_cutoff(dt: pd.Series, tz_mode: str) -> pd.Series:
    tz_norm = (tz_mode or "WIB").strip().upper()
    if dt.empty or tz_norm == "WIB":
        return dt

    adjusted = dt.copy()

    if tz_norm == "WITA":
        mask = adjusted.notna() & adjusted.dt.hour.eq(0)
        adjusted.loc[mask] = adjusted.loc[mask] - pd.Timedelta(days=1)
        return adjusted

    if tz_norm == "WIT":
        mask = adjusted.notna() & adjusted.dt.hour.isin([0, 1])
        adjusted.loc[mask] = adjusted.loc[mask] - pd.Timedelta(days=1)
        return adjusted

    return adjusted


def _parse_ticket_created_series(sr: pd.Series, tz_mode: str = "WIB") -> pd.Series:
    raw = sr.fillna("").astype(str).str.strip()
    dt = pd.to_datetime(raw, format="%d/%m/%Y %H:%M", errors="coerce")
    missing = dt.isna() & raw.ne("")

    if missing.any():
        dt_seconds = pd.to_datetime(raw[missing], format="%d/%m/%Y %H:%M:%S", errors="coerce")
        dt.loc[missing] = dt_seconds

    dt = _adjust_created_cutoff(dt, tz_mode)
    return dt.dt.normalize()


def _find_ticket_date_col(df: pd.DataFrame) -> Optional[str]:
    return (
        _find_col(df, ["Created", "Created Date", "Created At", "Created Time"])
        or _find_col(df, ["Action Date", "Action"])
    )


def _idr_fmt(val) -> str:
    if val is None:
        return "-"
    if isinstance(val, float) and np.isnan(val):
        return "-"
    try:
        n = float(val)
    except Exception:
        return str(val)
    neg = n < 0
    s = f"{abs(int(round(n))):,}".replace(",", ".")
    return f"({s})" if neg else s


def _style_right(df: pd.DataFrame):
    if df is None or getattr(df, "empty", True):
        return df

    cols = list(df.columns)

    def _last_label(col):
        return str(col[-1]) if isinstance(col, tuple) else str(col)

    def _is_no(col):
        return _last_label(col) == "NO"

    def _is_left(col):
        return _last_label(col) in ("TANGGAL", "Tanggal", "ORDER ID", "Order ID", "STATUS", "Status")

    no_cols = [c for c in cols if _is_no(c)]
    left_cols = [c for c in cols if _is_left(c)]
    right_cols = [c for c in cols if c not in set(no_cols + left_cols)]

    sty = df.style
    sty = sty.set_table_styles([{"selector": "th", "props": [("text-align", "center")]}])

    if no_cols:
        sty = sty.set_properties(subset=no_cols, **{"text-align": "center"})
    if left_cols:
        sty = sty.set_properties(subset=left_cols, **{"text-align": "left"})
    if right_cols:
        sty = sty.set_properties(subset=right_cols, **{"text-align": "right"})

    return sty


def _concat_files(files, reader: Callable = _read_any) -> pd.DataFrame:
    if not files:
        return pd.DataFrame()
    frames = []
    for f in files:
        df = reader(f)
        if not df.empty:
            df = _clean_columns(df)
            df["__source__"] = getattr(f, "name", "file")
            frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _expand_zip(files):
    if not files:
        return []
    out = []
    allow_ext = (".csv", ".xls", ".xlsx")
    for f in files:
        fname = (f.name or "").lower()
        if fname.endswith(".zip"):
            try:
                f.seek(0)
                with zipfile.ZipFile(f) as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        inner = info.filename
                        inner_lower = inner.lower()
                        if not inner_lower.endswith(allow_ext):
                            continue
                        if inner_lower.startswith("__macosx/") or inner_lower.endswith(".ds_store"):
                            continue
                        data = zf.read(info)
                        bio = io.BytesIO(data)
                        bio.name = f"{f.name}::{inner}"
                        out.append(bio)
            except Exception as e:
                st.warning(f"Gagal ekstrak ZIP {f.name}: {e}")
        else:
            out.append(f)
    return out


def _promote_header(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    def _is_header_row(sr: pd.Series) -> bool:
        vals = [str(v).strip().lower() for v in sr.fillna("")]
        keys = [
            "date",
            "tanggal",
            "transaction date",
            "tgl",
            "remark",
            "keterangan",
            "description",
            "deskripsi",
            "credit",
            "kredit",
            "cr",
            "amount",
            "jumlah",
        ]
        score = sum(any(k in v for k in keys) for v in vals)
        return score >= 2

    for r in (12, 13):
        if r < len(df) and _is_header_row(df.iloc[r]):
            cols = [str(x).strip() for x in df.iloc[r].tolist()]
            out = df.iloc[r + 1 :].copy()
            out.columns = cols
            out.columns = [str(c).strip().lstrip("\ufeff") for c in out.columns]
            return out

    scan_max = min(50, len(df))
    for r in range(scan_max):
        if _is_header_row(df.iloc[r]):
            cols = [str(x).strip() for x in df.iloc[r].tolist()]
            out = df.iloc[r + 1 :].copy()
            out.columns = cols
            out.columns = [str(c).strip().lstrip("\ufeff") for c in out.columns]
            return out

    return df


def _concat_rk_non(files) -> pd.DataFrame:
    if not files:
        return pd.DataFrame()
    frames = []
    for f in files:
        df = _read_any(f)
        if df.empty:
            continue
        df = _promote_header(df)
        if df.empty:
            continue
        df.columns = [str(c).strip().lstrip("\ufeff") for c in df.columns]
        df["__source__"] = getattr(f, "name", "file")
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _month_selector() -> Tuple[int, int]:
    from datetime import date

    today = date.today()
    years = list(range(today.year - 5, today.year + 2))
    months = [
        ("01", "Januari"),
        ("02", "Februari"),
        ("03", "Maret"),
        ("04", "April"),
        ("05", "Mei"),
        ("06", "Juni"),
        ("07", "Juli"),
        ("08", "Agustus"),
        ("09", "September"),
        ("10", "Oktober"),
        ("11", "November"),
        ("12", "Desember"),
    ]
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Tahun", years, index=years.index(today.year))
    with col2:
        month_label = st.selectbox("Bulan", months, index=int(today.strftime("%m")) - 1, format_func=lambda x: x[1])
        month = int(month_label[0])
    return year, month


# ---------- App ----------

st.set_page_config(page_title="Rekonsiliasi Tiket vs Settlement", layout="wide")
st.title("Rekonsiliasi: Tiket Detail vs Settlement Dana")

# --- FORCE ALIGNMENT: angka kanan, NO center, TANGGAL kiri (UI Streamlit) ---
st.markdown(
    """
    <style>
    div[data-testid="stDataFrame"] td { text-align: right !important; }
    div[data-testid="stDataFrame"] th { text-align: center !important; }
    div[data-testid="stDataFrame"] td:nth-child(1) { text-align: center !important; }
    div[data-testid="stDataFrame"] td:nth-child(2) { text-align: left !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

if "HASIL" not in st.session_state:
    st.session_state["HASIL"] = {}
if "LAST_PARAMS" not in st.session_state:
    st.session_state["LAST_PARAMS"] = None

with st.sidebar:
    st.header("1) Upload Sumber (multi-file)")
    tiket_files = st.file_uploader(
        "Tiket Detail (CSV/Excel .csv/.xls/.xlsx/.zip)",
        type=["csv", "xls", "xlsx", "zip"],
        accept_multiple_files=True,
    )
    settle_files = st.file_uploader(
        "Settlement Dana (CSV/Excel/.zip)",
        type=["csv", "xls", "xlsx", "zip"],
        accept_multiple_files=True,
    )
    st.divider()
    st.header("Rekening Koran (opsional, multi-file)")
    rk_bca_files = st.file_uploader(
        "Rekening Koran BCA (CSV/Excel/.zip)",
        type=["csv", "xls", "xlsx", "zip"],
        accept_multiple_files=True,
    )
    rk_non_files = st.file_uploader(
        "Rekening Koran Non BCA (CSV/Excel/.zip)",
        type=["csv", "xls", "xlsx", "zip"],
        accept_multiple_files=True,
    )

    st.header("2) Parameter Bulan & Tahun (WAJIB)")
    y, m = _month_selector()
    month_start = pd.Timestamp(y, m, 1)
    month_end = pd.Timestamp(y, m, calendar.monthrange(y, m)[1])

    st.header("3) Zona Waktu Tiket Detail")
    ticket_tz_mode = st.selectbox(
        "Penyesuaian kolom Created",
        options=["WIB", "WITA", "WIT"],
        index=0,
        help=(
            "WIB: tidak ada penyesuaian. "
            "WITA: Created jam 00:00:00-00:59:59 mundur 1 hari. "
            "WIT: Created jam 00:00:00-01:59:59 mundur 1 hari."
        ),
    )

    params_fingerprint = f"{y}-{m:02d}-{ticket_tz_mode}"
    last_params = st.session_state.get("LAST_PARAMS")
    if last_params is not None and last_params != params_fingerprint:
        st.session_state["HASIL"] = {}
    st.session_state["LAST_PARAMS"] = params_fingerprint

    st.caption(f"Periode dipakai: {month_start.date()} s/d {month_end.date()} | Zona waktu Created: {ticket_tz_mode}")

    go = st.button("Proses", type="primary", use_container_width=True)

    if st.session_state["HASIL"]:
        if st.button("Reset hasil (hapus tampilan tersimpan)", type="secondary", use_container_width=True):
            st.session_state["HASIL"] = {}
            st.rerun()

if go:
    tiket_inputs = _expand_zip(tiket_files)
    settle_inputs = _expand_zip(settle_files)
    rk_bca_inputs = _expand_zip(rk_bca_files)
    rk_non_inputs = _expand_zip(rk_non_files)

    tiket_df = _concat_files(tiket_inputs, reader=_read_tiket_detail_any)
    settle_df = _concat_files(settle_inputs)
    rk_bca_df = _concat_files(rk_bca_inputs)
    rk_non_df = _concat_rk_non(rk_non_inputs)

    if tiket_df.empty:
        st.error("Tiket Detail kosong / belum diupload.")
        st.stop()
    if settle_df.empty:
        st.error("Settlement Dana kosong / belum diupload.")
        st.stop()

    # ---------------------- Tiket Detail (TABEL 1) ----------------------
    t_created = _find_col(tiket_df, ["Created", "Created Date", "Created At", "Created Time"])
    t_date_action = _find_ticket_date_col(tiket_df)
    if t_date_action is None:
        st.error("Kolom tanggal 'Created' / 'Action Date' / 'Action' tidak ditemukan pada Tiket Detail.")
        st.stop()

    if t_created is not None:
        tiket_df["__ticket_date__"] = _parse_ticket_created_series(tiket_df[t_created], tz_mode=ticket_tz_mode)
        if tiket_df["__ticket_date__"].notna().any():
            t_date_action = "__ticket_date__"

    t_amt_tarif = _find_col(tiket_df, ["Tarif", "tarif"])
    if t_amt_tarif is None:
        st.error("Kolom nominal 'Tarif' tidak ditemukan pada Tiket Detail.")
        st.stop()

    t_stat = _find_col(tiket_df, ["St Bayar", "Status Bayar", "status", "status bayar"])
    t_bank = _find_col(tiket_df, ["Bank", "Payment Channel", "channel", "payment method"])
    if t_stat is None or t_bank is None:
        st.error("Kolom 'St Bayar' atau 'Bank' tidak ditemukan pada Tiket Detail.")
        st.stop()

    # ---------------------- Mapping Settlement (utama/legacy) ----------------------
    s_date_legacy = _find_col(settle_df, ["Transaction Date", "Tanggal Transaksi", "Tanggal"])
    s_amt_legacy = _find_col(settle_df, ["Settlement Amount", "Amount", "Nominal", "Jumlah"])
    if s_amt_legacy is None and not settle_df.empty and len(settle_df.columns) >= 12:
        s_amt_legacy = settle_df.columns[11]
    if s_date_legacy is None:
        s_date_legacy = _find_col(settle_df, ["Settlement Date", "Tanggal Settlement", "Settle Date", "Tanggal", "Setle Date"])
        if s_date_legacy is None and not settle_df.empty and len(settle_df.columns) >= 5:
            s_date_legacy = settle_df.columns[4]

    if s_date_legacy is None or s_amt_legacy is None:
        st.error("Kolom wajib Settlement Dana tidak ditemukan (tanggal/amount).")
        st.stop()

    # Untuk BCA/Non-BCA (pakai E/L/P)
    s_date_E = _find_col(settle_df, ["Settlement Date", "Tanggal Settlement", "Settle Date", "Tanggal"])
    s_amt_L = _find_col(settle_df, ["Settlement Amount", "Amount", "Nominal", "Jumlah"])
    if s_amt_L is None and not settle_df.empty and len(settle_df.columns) >= 12:
        s_amt_L = settle_df.columns[11]
    s_prod_P = _find_col(settle_df, ["Product Name", "Produk", "Nama Produk"])
    if s_date_E is None and not settle_df.empty and len(settle_df.columns) >= 5:
        s_date_E = settle_df.columns[4]
    if s_prod_P is None and not settle_df.empty and len(settle_df.columns) >= 16:
        s_prod_P = settle_df.columns[15]

    # ------------------  TABEL 1: TIKET DETAIL ESPAY (PAID ONLY) -------------------
    td = tiket_df.copy()
    if t_date_action == "__ticket_date__":
        td[t_date_action] = pd.to_datetime(td[t_date_action], errors="coerce")
    else:
        if t_created is not None and t_date_action != t_created:
            td = _fill_action_from_created(td, t_date_action, t_created)
        td[t_date_action] = pd.to_datetime(td[t_date_action].apply(_to_date), errors="coerce")
    td = td[~td[t_date_action].isna()]

    td_bank_norm = td[t_bank].apply(_norm_str)
    td_stat_norm = td[t_stat].apply(_norm_str)
    td = td[td_bank_norm.eq("espay") & td_stat_norm.eq("paid")]
    td = td[(td[t_date_action] >= month_start) & (td[t_date_action] <= month_end)]

    td[t_amt_tarif] = _to_num(td[t_amt_tarif])

    # FIX: ABAIKAN DUPLICATE UNTUK TIKET DETAIL (jangan drop_duplicates)
    tiket_by_date = td.groupby(td[t_date_action].dt.date, dropna=True)[t_amt_tarif].sum()

    # ------------------  Settlement Dana (utama/legacy) ------------------
    sd_main = settle_df.copy()
    sd_main[s_date_legacy] = sd_main[s_date_legacy].apply(_to_date)
    sd_main = sd_main[~sd_main[s_date_legacy].isna()]
    sd_main = sd_main[(sd_main[s_date_legacy] >= month_start) & (sd_main[s_date_legacy] <= month_end)]
    sd_main[s_amt_legacy] = _to_num(sd_main[s_amt_legacy])
    settle_by_date_total = sd_main.groupby(sd_main[s_date_legacy].dt.date, dropna=True)[s_amt_legacy].sum()

    # --- Settlement BCA / Non BCA dari E/L/P (untuk split) ---
    bca_series = pd.Series(dtype=float)
    non_bca_series = pd.Series(dtype=float)
    if not settle_df.empty and s_date_E and s_amt_L and s_prod_P:
        sd_bca = settle_df.copy()
        sd_bca[s_date_E] = sd_bca[s_date_E].apply(_to_date)
        sd_bca = sd_bca[~sd_bca[s_date_E].isna()]
        sd_bca = sd_bca[(sd_bca[s_date_E] >= month_start) & (sd_bca[s_date_E] <= month_end)]
        sd_bca[s_amt_L] = _to_num(sd_bca[s_amt_L])
        prod_norm = sd_bca[s_prod_P]
        bca_mask = prod_norm.apply(_is_bca_va_product)
        settle_by_date_bca = sd_bca.loc[bca_mask].groupby(sd_bca[s_date_E].dt.date, dropna=True)[s_amt_L].sum()
        settle_by_date_non_bca = sd_bca.loc[~bca_mask].groupby(sd_bca[s_date_E].dt.date, dropna=True)[s_amt_L].sum()
        bca_series = settle_by_date_bca
        non_bca_series = settle_by_date_non_bca

    # --- RK: Uang Masuk BCA ---
    uang_masuk_bca = pd.Series(dtype=float)
    if not rk_bca_df.empty:
        rk_tgl_bca = _find_col(rk_bca_df, ["Tanggal", "Date", "Tgl", "Transaction Date"])
        rk_amt_bca = _find_col(rk_bca_df, ["mutasi", "amount", "kredit", "credit", "cr"])
        rk_ket_bca = _find_col(rk_bca_df, ["Keterangan", "Remark", "Deskripsi", "Description"])
        if rk_tgl_bca and rk_amt_bca and rk_ket_bca:
            bca = rk_bca_df.copy()
            bca[rk_tgl_bca] = bca[rk_tgl_bca].apply(_to_date)
            bca = bca[~bca[rk_tgl_bca].isna()]
            bca = bca[(bca[rk_tgl_bca] >= month_start) & (bca[rk_tgl_bca] <= month_end)]
            ket_norm = bca[rk_ket_bca].astype(str).str.strip().str.lower()
            bca = bca[ket_norm.str.contains("mrc", na=False)]
            bca[rk_amt_bca] = _to_num(bca[rk_amt_bca])
            uang_masuk_bca = bca.groupby(bca[rk_tgl_bca].dt.date, dropna=True)[rk_amt_bca].sum()

    # --- RK: Uang Masuk NON BCA ---
    uang_masuk_non = pd.Series(dtype=float)
    if not rk_non_df.empty:
        rk_tgl_non = _find_col(rk_non_df, ["Date", "Tanggal", "Transaction Date", "Tgl"])
        rk_amt_non = _find_col(rk_non_df, ["credit", "kredit", "cr", "amount"])
        rk_rem_non = _find_col(rk_non_df, ["Remark", "Keterangan", "Description", "Deskripsi"])
        if rk_tgl_non and rk_amt_non and rk_rem_non:
            nb = rk_non_df.copy()
            nb[rk_tgl_non] = nb[rk_tgl_non].apply(_to_date)
            nb = nb[~nb[rk_tgl_non].isna()]
            nb = nb[(nb[rk_tgl_non] >= month_start) & (nb[rk_tgl_non] <= month_end)]
            rem_norm = nb[rk_rem_non].astype(str).str.strip().str.lower()
            nb = nb[rem_norm.str.contains("mrc", na=False)]
            nb[rk_amt_non] = _to_num(nb[rk_amt_non])
            uang_masuk_non = nb.groupby(nb[rk_tgl_non].dt.date, dropna=True)[rk_amt_non].sum()

    # --- Index tanggal (1..akhir bulan) ---
    idx = pd.Index(pd.date_range(month_start, month_end, freq="D").date, name="Tanggal")
    tiket_series = tiket_by_date.reindex(idx, fill_value=0.0)
    settle_series = settle_by_date_total.reindex(idx, fill_value=0.0)
    bca_series = bca_series.reindex(idx, fill_value=0.0)
    non_bca_series = non_bca_series.reindex(idx, fill_value=0.0)
    total_settle_ser = (bca_series + non_bca_series).reindex(idx, fill_value=0.0)
    uang_masuk_bca_ser = uang_masuk_bca.reindex(idx, fill_value=0.0)
    uang_masuk_non_ser = uang_masuk_non.reindex(idx, fill_value=0.0)
    total_uang_masuk_ser = (uang_masuk_bca_ser + uang_masuk_non_ser).reindex(idx, fill_value=0.0)

    final = pd.DataFrame(index=idx)
    final["TIKET DETAIL ESPAY"] = tiket_series.values
    final["SETTLEMENT DANA ESPAY"] = settle_series.values
    final["SELISIH TIKET DETAIL - SETTLEMENT"] = final["TIKET DETAIL ESPAY"] - final["SETTLEMENT DANA ESPAY"]
    final["SETTLEMENT BCA"] = bca_series.values
    final["SETTLEMENT NON BCA"] = non_bca_series.values
    final["TOTAL SETTLEMENT"] = (bca_series.values + non_bca_series.values)
    final["UANG MASUK BCA"] = uang_masuk_bca_ser.values
    final["UANG MASUK NON BCA"] = uang_masuk_non_ser.values
    final["TOTAL UANG MASUK"] = total_uang_masuk_ser.values
    final["SELISIH SETTLEMENT - UANG MASUK"] = final["TOTAL SETTLEMENT"] - final["TOTAL UANG MASUK"]

    view = final.reset_index()
    idx_col_name = view.columns[0]
    view = view.rename(columns={idx_col_name: "TANGGAL"})
    view.insert(0, "NO", range(1, len(view) + 1))

    total_row = pd.DataFrame([{
        "NO": "",
        "TANGGAL": "TOTAL",
        "TIKET DETAIL ESPAY": final["TIKET DETAIL ESPAY"].sum(),
        "SETTLEMENT DANA ESPAY": final["SETTLEMENT DANA ESPAY"].sum(),
        "SELISIH TIKET DETAIL - SETTLEMENT": final["SELISIH TIKET DETAIL - SETTLEMENT"].sum(),
        "SETTLEMENT BCA": final["SETTLEMENT BCA"].sum(),
        "SETTLEMENT NON BCA": final["SETTLEMENT NON BCA"].sum(),
        "TOTAL SETTLEMENT": final["TOTAL SETTLEMENT"].sum(),
        "UANG MASUK BCA": final["UANG MASUK BCA"].sum(),
        "UANG MASUK NON BCA": final["UANG MASUK NON BCA"].sum(),
        "TOTAL UANG MASUK": final["TOTAL UANG MASUK"].sum(),
        "SELISIH SETTLEMENT - UANG MASUK": final["SELISIH SETTLEMENT - UANG MASUK"].sum(),
    }])

    view_total = pd.concat([view, total_row], ignore_index=True)

    ordered_cols = [
        "NO", "TANGGAL",
        "TIKET DETAIL ESPAY", "SETTLEMENT DANA ESPAY", "SELISIH TIKET DETAIL - SETTLEMENT",
        "SETTLEMENT BCA", "SETTLEMENT NON BCA", "TOTAL SETTLEMENT",
        "UANG MASUK BCA", "UANG MASUK NON BCA", "TOTAL UANG MASUK",
        "SELISIH SETTLEMENT - UANG MASUK",
    ]
    view_total = view_total.loc[:, ordered_cols]

    fmt = view_total.copy()
    for c in ordered_cols:
        if c in ("NO", "TANGGAL"):
            continue
        fmt[c] = fmt[c].apply(_idr_fmt)

    from openpyxl.styles import Alignment, Font

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        view_total.to_excel(xw, index=False, sheet_name="Rekonsiliasi")
        fmt.to_excel(xw, index=False, sheet_name="Rekonsiliasi_View")

        wb = xw.book
        ws = wb["Rekonsiliasi_View"]
        ws.insert_rows(1)

        sub_headers = [
            "NO", "TANGGAL",
            "TIKET DETAIL ESPAY", "SETTLEMENT DANA ESPAY", "SELISIH TIKET DETAIL - SETTLEMENT",
            "BCA", "NON BCA", "TOTAL SETTLEMENT",
            "BCA", "NON BCA", "TOTAL UANG MASUK",
            "SELISIH SETTLEMENT - UANG MASUK",
        ]
        top_headers = [
            "NO", "TANGGAL",
            "TIKET DETAIL ESPAY", "SETTLEMENT DANA ESPAY", "SELISIH TIKET DETAIL - SETTLEMENT",
            "SETTLEMENT", "SETTLEMENT", "TOTAL SETTLEMENT",
            "UANG MASUK", "UANG MASUK", "TOTAL UANG MASUK",
            "SELISIH SETTLEMENT - UANG MASUK",
        ]

        for col_idx, (top, sub) in enumerate(zip(top_headers, sub_headers), start=1):
            ws.cell(row=1, column=col_idx, value=top)
            ws.cell(row=2, column=col_idx, value=sub)

        ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)

        max_col = ws.max_column
        for c in range(1, max_col + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=2, column=c).font = Font(bold=True)
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22

    # ======================================================================
    # ===========  TABEL: DETAIL TIKET (GO SHOW × SUB-KATEGORI)  ===========
    # ======================================================================

    def _col_by_letter_local(df: pd.DataFrame, letters: str) -> Optional[str]:
        if df is None or df.empty:
            return None
        s = letters.strip().upper()
        if not s:
            return None
        n = 0
        for ch in s:
            if not ("A" <= ch <= "Z"):
                return None
            n = n * 26 + (ord(ch) - ord("A") + 1)
        idx0 = n - 1
        return df.columns[idx0] if 0 <= idx0 < len(df.columns) else None

    type_main_col = _find_col(tiket_df, ["Type", "Tipe", "Jenis"]) or _col_by_letter_local(tiket_df, "B")
    bank_col = _find_col(tiket_df, ["Bank", "Payment Channel", "channel", "payment method"]) or _col_by_letter_local(tiket_df, "I")
    type_sub_col = (
        _find_col(
            tiket_df,
            [
                "Payment Type", "Channel Type", "Transaction Type", "Sub Type",
                "Tipe", "Tipe Pembayaran", "Jenis Pembayaran", "Kategori", "Metode", "Product Type",
            ],
        )
        or _col_by_letter_local(tiket_df, "J")
    )
    date_col = "__ticket_date__" if "__ticket_date__" in tiket_df.columns else (_find_ticket_date_col(tiket_df) or _col_by_letter_local(tiket_df, "AG"))
    tarif_col = _find_col(tiket_df, ["Tarif", "tarif"]) or _col_by_letter_local(tiket_df, "Y")
    status_col = _find_col(tiket_df, ["St Bayar", "Status Bayar", "status", "status bayar"])

    required_missing = [
        n for n, c in [
            ("TYPE (kolom B)", type_main_col),
            ("BANK (kolom I)", bank_col),
            ("TIPE / SUB-TIPE (kolom J)", type_sub_col),
            ("CREATED / ACTION DATE (kolom AG fallback)", date_col),
            ("TARIF (kolom Y)", tarif_col),
            ("ST BAYAR / STATUS BAYAR", status_col),
        ]
        if c is None
    ]

    df2_fmt_mi = None
    if required_missing:
        st.warning("Kolom wajib untuk tabel 'Detail Tiket (GO SHOW/ONLINE) [PAID]' belum lengkap: " + ", ".join(required_missing))
    else:
        tix = tiket_df.copy()
        if date_col == "__ticket_date__":
            tix[date_col] = pd.to_datetime(tix[date_col], errors="coerce")
        else:
            if t_created is not None and date_col != t_created:
                tix = _fill_action_from_created(tix, date_col, t_created)
            tix[date_col] = pd.to_datetime(tix[date_col].apply(_to_date), errors="coerce")
        tix = tix[~tix[date_col].isna()]
        tix = tix[(tix[date_col] >= month_start) & (tix[date_col] <= month_end)]

        stat_norm = tix[status_col].apply(_norm_str)
        tix = tix[stat_norm.eq("paid") | stat_norm.str.contains(r"\bpaid\b", na=False)]

        main_norm_all = tix[type_main_col].apply(_norm_str)
        sub_norm_all = tix[type_sub_col].apply(_norm_str)
        bank_norm_all = tix[bank_col].apply(_norm_bank)

        tix[tarif_col] = _to_num(tix[tarif_col])

        m_go_show = (main_norm_all == "go show") | main_norm_all.str.contains(r"\bgo\s*show\b", na=False)
        m_online = (main_norm_all == "online") | main_norm_all.str.contains(r"\bonline\b", na=False)

        m_prepaid_all = (sub_norm_all == "prepaid") | sub_norm_all.str.contains(r"\bprepaid\b", na=False)
        m_emoney_all = (sub_norm_all == "e-money") | sub_norm_all.str.contains(r"\be[-\s]*money\b|\bemoney\b", na=False)
        m_varetail_all = sub_norm_all.str.contains(r"virtual\s*account", na=False) & sub_norm_all.str.contains(r"gerai|retail", na=False)
        m_cash_all = (sub_norm_all == "cash") | sub_norm_all.str.contains(r"\bcash\b", na=False)

        # TAMBAHAN: TRANSFER (kolom J mengandung "transfer")
        m_tf_all = sub_norm_all.str.contains(r"\btransfer\b", na=False)

        idx2 = pd.Index(pd.date_range(month_start, month_end, freq="D").date, name="Tanggal")

        # ---------------- GO SHOW existing ----------------
        s_gs_prepaid_bca = tix.loc[m_go_show & m_prepaid_all & bank_norm_all.eq("bca")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_prepaid_bri = tix.loc[m_go_show & m_prepaid_all & bank_norm_all.eq("bri")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_prepaid_bni = tix.loc[m_go_show & m_prepaid_all & bank_norm_all.eq("bni")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_prepaid_mandiri = tix.loc[m_go_show & m_prepaid_all & bank_norm_all.eq("mandiri")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_emoney_espay = tix.loc[m_go_show & m_emoney_all & bank_norm_all.eq("espay")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_varetail_espay = tix.loc[m_go_show & m_varetail_all & bank_norm_all.eq("espay")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_cash_asdp = tix.loc[m_go_show & m_cash_all & bank_norm_all.eq("asdp")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()

        # ---------------- GO SHOW TRANSFER (BARU) ----------------
        # PT.POS variasi: "pt.pos", "pt pos", "ptpos", "pos"
        m_bank_ptpos = bank_norm_all.str.contains(r"\b(pt\.?\s*pos|ptpos|pos)\b", na=False)

        s_gs_tf_ptpos = tix.loc[m_go_show & m_tf_all & m_bank_ptpos].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_tf_bri = tix.loc[m_go_show & m_tf_all & bank_norm_all.eq("bri")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_tf_bni = tix.loc[m_go_show & m_tf_all & bank_norm_all.eq("bni")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_tf_mandiri = tix.loc[m_go_show & m_tf_all & bank_norm_all.eq("mandiri")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_gs_tf_bca = tix.loc[m_go_show & m_tf_all & bank_norm_all.eq("bca")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()

        go_show_cols = {
            "PREPAID - BCA": s_gs_prepaid_bca.reindex(idx2, fill_value=0.0),
            "PREPAID - BRI": s_gs_prepaid_bri.reindex(idx2, fill_value=0.0),
            "PREPAID - BNI": s_gs_prepaid_bni.reindex(idx2, fill_value=0.0),
            "PREPAID - MANDIRI": s_gs_prepaid_mandiri.reindex(idx2, fill_value=0.0),

            # KOLOM BARU (TRANSFER) MASUK GRUP GO SHOW
            "TF - PT.POS": s_gs_tf_ptpos.reindex(idx2, fill_value=0.0),
            "TF - BRI": s_gs_tf_bri.reindex(idx2, fill_value=0.0),
            "TF - BNI": s_gs_tf_bni.reindex(idx2, fill_value=0.0),
            "TF - MANDIRI": s_gs_tf_mandiri.reindex(idx2, fill_value=0.0),
            "TF - BCA": s_gs_tf_bca.reindex(idx2, fill_value=0.0),

            "E-MONEY - ESPAY": s_gs_emoney_espay.reindex(idx2, fill_value=0.0),
            "VIRTUAL ACCOUNT DAN GERAI RETAIL - ESPAY": s_gs_varetail_espay.reindex(idx2, fill_value=0.0),
            "CASH - ASDP": s_gs_cash_asdp.reindex(idx2, fill_value=0.0),
        }

        # ---------------- ONLINE existing ----------------
        m_emoney_on = (sub_norm_all == "e-money") | sub_norm_all.str.contains(r"\be[-\s]*money\b|\bemoney\b", na=False)
        m_varetail_on = sub_norm_all.str.contains(r"virtual\s*account", na=False) & sub_norm_all.str.contains(r"gerai|retail", na=False)
        m_cash_on = (sub_norm_all == "cash") | sub_norm_all.str.contains(r"\bcash\b", na=False)
        m_bank_espay_on = bank_norm_all.eq("espay")

        s_on_emoney_espay = tix.loc[m_online & m_bank_espay_on & m_emoney_on].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_on_varetail_espay = tix.loc[m_online & m_bank_espay_on & m_varetail_on].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()
        s_on_cash_asdp = tix.loc[m_online & m_cash_on & bank_norm_all.eq("asdp")].groupby(tix[date_col].dt.date, dropna=True)[tarif_col].sum()

        online_cols = {
            "E-MONEY - ESPAY": s_on_emoney_espay.reindex(idx2, fill_value=0.0),
            "VIRTUAL ACCOUNT & GERAI RETAIL - ESPAY": s_on_varetail_espay.reindex(idx2, fill_value=0.0),
            "CASH - ASDP": s_on_cash_asdp.reindex(idx2, fill_value=0.0),
        }

        detail_mix = pd.DataFrame(index=idx2)

        for k, ser in go_show_cols.items():
            detail_mix[f"GS|{k}"] = ser.values
        detail_mix["GS|SUBTOTAL"] = pd.DataFrame(go_show_cols, index=idx2).sum(axis=1).values

        for k, ser in online_cols.items():
            detail_mix[f"ON|{k}"] = ser.values
        detail_mix["ON|SUBTOTAL"] = pd.DataFrame(online_cols, index=idx2).sum(axis=1).values

        detail_mix["GT|GRAND TOTAL"] = detail_mix["GS|SUBTOTAL"] + detail_mix["ON|SUBTOTAL"]

        df2 = detail_mix.reset_index()
        df2.insert(0, "NO", range(1, len(df2) + 1))

        total_row2 = {"NO": "", "Tanggal": "TOTAL"}
        for k in detail_mix.columns:
            total_row2[k] = float(detail_mix[k].sum())
        df2 = pd.concat([df2, pd.DataFrame([total_row2])], ignore_index=True)

        df2_fmt = df2.copy()
        for c in df2_fmt.columns:
            if c in ("NO", "Tanggal"):
                continue
            df2_fmt[c] = df2_fmt[c].apply(_idr_fmt)

        def _strip_prefix(col_name: str) -> tuple[str, str]:
            if col_name.startswith("GS|"):
                return ("GO SHOW", col_name[3:])
            if col_name.startswith("ON|"):
                return ("ONLINE", col_name[3:])
            if col_name.startswith("GT|"):
                return ("GRAND TOTAL", "")
            return ("", col_name)

        ordered_keys = [k for k in detail_mix.columns if k.startswith("GS|") and k != "GS|SUBTOTAL"] + ["GS|SUBTOTAL"]
        ordered_keys += [k for k in detail_mix.columns if k.startswith("ON|") and k != "ON|SUBTOTAL"] + ["ON|SUBTOTAL"]
        ordered_keys += ["GT|GRAND TOTAL"]

        df2_fmt = df2_fmt[["NO", "Tanggal"] + ordered_keys]
        top = [("", "NO"), ("", "Tanggal")] + [_strip_prefix(k) for k in ordered_keys]
        df2_fmt_mi = df2_fmt.copy()
        df2_fmt_mi.columns = pd.MultiIndex.from_tuples(top)

    # ======================================================================
    # ===================  TABEL: DETAIL SETTLEMENT REPORT  =================
    # ======================================================================

    detail_settle_table = None
    detail_settle_excel_bytes = None

    s_order = _find_col(settle_df, ["Order ID", "OrderId", "Order Number", "Order No", "OrderID", "order id"])
    miss = [n for n, c in [
        ("Settlement Date (E)", s_date_E),
        ("Settlement Amount (L)", s_amt_L),
        ("Product Name (P)", s_prod_P),
        ("Order ID", s_order),
    ] if c is None]

    if miss:
        st.warning("Kolom untuk 'DETAIL SETTLEMENT REPORT' belum lengkap: " + ", ".join(miss))
    else:
        sd = settle_df.copy()
        sd[s_date_E] = sd[s_date_E].apply(_to_date)
        sd = sd[~sd[s_date_E].isna()]
        sd = sd[(sd[s_date_E] >= month_start) & (sd[s_date_E] <= month_end)]

        sd[s_amt_L] = _to_num(sd[s_amt_L])
        prod_raw = sd[s_prod_P]
        prod_norm = prod_raw.astype(str).str.strip().str.casefold()
        order_norm = sd[s_order].astype(str).str.strip().str.casefold()

        go_show_mask = order_norm.str.endswith("_ord") | (~order_norm.str.startswith("ord") & order_norm.str.endswith("ord"))
        online_mask = order_norm.str.startswith("ord")

        is_bca_va = prod_raw.apply(_is_bca_va_product)
        has_va = prod_norm.str.contains("va", na=False) | is_bca_va
        non_bca_va = has_va & ~is_bca_va
        is_emoney = ~has_va

        gs_va_bca = sd.loc[go_show_mask & is_bca_va].groupby(sd[s_date_E].dt.date, dropna=True)[s_amt_L].sum()
        gs_va_nonbca = sd.loc[go_show_mask & non_bca_va].groupby(sd[s_date_E].dt.date, dropna=True)[s_amt_L].sum()
        gs_emoney = sd.loc[go_show_mask & is_emoney].groupby(sd[s_date_E].dt.date, dropna=True)[s_amt_L].sum()

        on_va_bca = sd.loc[online_mask & is_bca_va].groupby(sd[s_date_E].dt.date, dropna=True)[s_amt_L].sum()
        on_va_nonbca = sd.loc[online_mask & non_bca_va].groupby(sd[s_date_E].dt.date, dropna=True)[s_amt_L].sum()
        on_emoney = sd.loc[online_mask & is_emoney].groupby(sd[s_date_E].dt.date, dropna=True)[s_amt_L].sum()

        idx_set = pd.Index(pd.date_range(month_start, month_end, freq="D").date, name="Tanggal")
        detail_settle = pd.DataFrame(index=idx_set)

        detail_settle["GS|VIRTUAL ACCOUNT - BCA"] = gs_va_bca.reindex(idx_set, fill_value=0.0).values
        detail_settle["GS|VIRTUAL ACCOUNT - NON BCA"] = gs_va_nonbca.reindex(idx_set, fill_value=0.0).values
        detail_settle["GS|E-MONEY"] = gs_emoney.reindex(idx_set, fill_value=0.0).values

        detail_settle["ON|VIRTUAL ACCOUNT - BCA"] = on_va_bca.reindex(idx_set, fill_value=0.0).values
        detail_settle["ON|VIRTUAL ACCOUNT - NON BCA"] = on_va_nonbca.reindex(idx_set, fill_value=0.0).values
        detail_settle["ON|E-MONEY"] = on_emoney.reindex(idx_set, fill_value=0.0).values

        detail_settle["GS|Total Settlement (Go Show)"] = (
            detail_settle["GS|VIRTUAL ACCOUNT - BCA"]
            + detail_settle["GS|VIRTUAL ACCOUNT - NON BCA"]
            + detail_settle["GS|E-MONEY"]
        )
        detail_settle["ON|Total Settlement (Online)"] = (
            detail_settle["ON|VIRTUAL ACCOUNT - BCA"]
            + detail_settle["ON|VIRTUAL ACCOUNT - NON BCA"]
            + detail_settle["ON|E-MONEY"]
        )

        detail_settle = detail_settle[
            [
                "GS|VIRTUAL ACCOUNT - BCA",
                "GS|VIRTUAL ACCOUNT - NON BCA",
                "GS|E-MONEY",
                "GS|Total Settlement (Go Show)",
                "ON|VIRTUAL ACCOUNT - BCA",
                "ON|VIRTUAL ACCOUNT - NON BCA",
                "ON|E-MONEY",
                "ON|Total Settlement (Online)",
            ]
        ]

        df3 = detail_settle.reset_index()
        df3.insert(0, "NO", range(1, len(df3) + 1))

        total_row3 = {"NO": "", "Tanggal": "TOTAL"}
        for k in detail_settle.columns:
            total_row3[k] = float(detail_settle[k].sum())
        df3 = pd.concat([df3, pd.DataFrame([total_row3])], ignore_index=True)

        df3_fmt = df3.copy()
        for c in df3_fmt.columns:
            if c in ("NO", "Tanggal"):
                continue
            df3_fmt[c] = df3_fmt[c].apply(_idr_fmt)

        def _split_head(col_name: str) -> tuple[str, str]:
            if col_name.startswith("GS|"):
                return ("GO SHOW", col_name[3:])
            if col_name.startswith("ON|"):
                return ("ONLINE", col_name[3:])
            return ("", col_name)

        ordered = list(detail_settle.columns)
        df3_fmt = df3_fmt[["NO", "Tanggal"] + ordered]
        top3 = [("", "NO"), ("", "Tanggal")] + [_split_head(k) for k in ordered]
        df3_fmt_mi = df3_fmt.copy()
        df3_fmt_mi.columns = pd.MultiIndex.from_tuples(top3)
        detail_settle_table = df3_fmt_mi

        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        bio_settle = io.BytesIO()
        with pd.ExcelWriter(bio_settle, engine="openpyxl") as xw3:
            df3.to_excel(xw3, index=False, sheet_name="Detail_Settlement")

            wsname3 = "Detail_Settlement_View"
            df3_fmt.to_excel(xw3, index=False, header=False, sheet_name=wsname3, startrow=2)
            wb3 = xw3.book
            ws3 = wb3[wsname3]

            cols3 = list(df3.columns)
            top_headers = []
            sub_headers = []
            for c in cols3:
                if c in ("NO", "Tanggal"):
                    top_headers.append("")
                    sub_headers.append(c)
                elif str(c).startswith("GS|"):
                    top_headers.append("GO SHOW")
                    sub_headers.append(str(c)[3:])
                elif str(c).startswith("ON|"):
                    top_headers.append("ONLINE")
                    sub_headers.append(str(c)[3:])
                else:
                    top_headers.append("")
                    sub_headers.append(str(c))

            for j, (top, sub) in enumerate(zip(top_headers, sub_headers), start=1):
                ws3.cell(row=1, column=j, value=top)
                ws3.cell(row=2, column=j, value=sub)

            def _merge_same_run(labels, row_idx):
                start = 0
                while start < len(labels):
                    end = start
                    while end + 1 < len(labels) and labels[end + 1] == labels[start]:
                        end += 1
                    label = labels[start]
                    if label not in ("", None) and end >= start:
                        ws3.merge_cells(start_row=row_idx, start_column=start + 1, end_row=row_idx, end_column=end + 1)
                    start = end + 1

            _merge_same_run(top_headers, row_idx=1)

            for j, top in enumerate(top_headers, start=1):
                if top == "":
                    ws3.merge_cells(start_row=1, start_column=j, end_row=2, end_column=j)

            max_col = len(cols3)
            for ccol in range(1, max_col + 1):
                ws3.cell(row=1, column=ccol).font = Font(bold=True)
                ws3.cell(row=2, column=ccol).font = Font(bold=True)
                ws3.cell(row=1, column=ccol).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws3.cell(row=2, column=ccol).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws3.row_dimensions[1].height = 22
            ws3.row_dimensions[2].height = 22

            sample_rows = min(50, df3_fmt.shape[0])
            for idx_col, col_name in enumerate(cols3, start=1):
                max_len = max(len(str(col_name)), len(str(sub_headers[idx_col - 1])), len(str(top_headers[idx_col - 1])))
                for r in range(3, 3 + sample_rows):
                    v = ws3.cell(row=r, column=idx_col).value
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
                ws3.column_dimensions[get_column_letter(idx_col)].width = min(max(10, max_len + 2), 45)

        detail_settle_excel_bytes = bio_settle.getvalue()


    # ======================================================================
    # ========================  RINCIAN SELISIH ORDER ID  ===================
    # ======================================================================

    rincian_selisih_table = None
    rincian_selisih_excel_bytes = None

    t_order = _find_col(tiket_df, ["Order ID", "OrderId", "Order No", "Order Number", "OrderID"])
    s_order_rincian = _find_col(settle_df, ["Order ID", "OrderId", "Order Number", "Order No", "OrderID", "order id"])
    s_date_transaction = _find_col(settle_df, ["Transaction Date", "Tanggal Transaksi"])

    miss_rincian = [n for n, c in [
        ("Order ID Tiket Detail", t_order),
        ("Created / Action Date / Action", t_date_action),
        ("St Bayar / Status Bayar", t_stat),
        ("Bank", t_bank),
        ("Tarif", t_amt_tarif),
        ("Order ID Settlement Dana", s_order_rincian),
        ("Transaction Date", s_date_transaction),
        ("Settlement Amount", s_amt_legacy),
    ] if c is None]

    if miss_rincian:
        st.warning("Kolom untuk 'RINCIAN SELISIH' belum lengkap: " + ", ".join(miss_rincian))
    else:
        td_gap = tiket_df.copy()
        if t_date_action == "__ticket_date__":
            td_gap[t_date_action] = pd.to_datetime(td_gap[t_date_action], errors="coerce")
        else:
            if t_created is not None and t_date_action != t_created:
                td_gap = _fill_action_from_created(td_gap, t_date_action, t_created)
            td_gap[t_date_action] = pd.to_datetime(td_gap[t_date_action].apply(_to_date), errors="coerce")
        td_gap = td_gap[~td_gap[t_date_action].isna()]
        td_gap = td_gap[(td_gap[t_date_action] >= month_start) & (td_gap[t_date_action] <= month_end)]

        bank_mask = td_gap[t_bank].apply(_norm_str).str.contains("espay", na=False)
        paid_mask = td_gap[t_stat].apply(_norm_str).eq("paid")
        td_gap = td_gap[bank_mask & paid_mask].copy()

        td_gap[t_amt_tarif] = _to_num(td_gap[t_amt_tarif])
        td_gap["__order_key__"] = td_gap[t_order].apply(_norm_order_id)
        td_gap["__order_display__"] = td_gap[t_order].apply(lambda x: "" if pd.isna(x) else str(x).strip())
        td_gap = td_gap[td_gap["__order_key__"].ne("")]

        tiket_order_cmp = td_gap.groupby("__order_key__", as_index=False).agg(
            ORDER_ID_TIKET=("__order_display__", _first_nonempty_text),
            TARIF_TIKET_DETAIL=(t_amt_tarif, "sum"),
        )

        sd_gap = settle_df.copy()
        sd_gap[s_date_transaction] = pd.to_datetime(sd_gap[s_date_transaction].apply(_to_date), errors="coerce")
        sd_gap = sd_gap[~sd_gap[s_date_transaction].isna()]
        sd_gap = sd_gap[(sd_gap[s_date_transaction] >= month_start) & (sd_gap[s_date_transaction] <= month_end)]

        sd_gap[s_amt_legacy] = _to_num(sd_gap[s_amt_legacy])
        sd_gap["__order_key__"] = sd_gap[s_order_rincian].apply(_norm_order_id)
        sd_gap["__order_display__"] = sd_gap[s_order_rincian].apply(lambda x: "" if pd.isna(x) else str(x).strip())
        sd_gap = sd_gap[sd_gap["__order_key__"].ne("")]

        settle_order_cmp = sd_gap.groupby("__order_key__", as_index=False).agg(
            ORDER_ID_SETTLEMENT=("__order_display__", _first_nonempty_text),
            SETTLEMENT_AMOUNT=(s_amt_legacy, "sum"),
        )

        rincian_cmp = tiket_order_cmp.merge(settle_order_cmp, on="__order_key__", how="outer")
        rincian_cmp["ORDER ID"] = rincian_cmp["ORDER_ID_TIKET"].fillna(rincian_cmp["ORDER_ID_SETTLEMENT"])
        rincian_cmp["TARIF TIKET DETAIL"] = rincian_cmp["TARIF_TIKET_DETAIL"].fillna(0.0)
        rincian_cmp["SETTLEMENT AMOUNT"] = rincian_cmp["SETTLEMENT_AMOUNT"].fillna(0.0)
        rincian_cmp["SELISIH"] = rincian_cmp["TARIF TIKET DETAIL"] - rincian_cmp["SETTLEMENT AMOUNT"]

        rincian_cmp["STATUS"] = np.select(
            [
                rincian_cmp["TARIF TIKET DETAIL"].eq(0) & rincian_cmp["SETTLEMENT AMOUNT"].ne(0),
                rincian_cmp["TARIF TIKET DETAIL"].ne(0) & rincian_cmp["SETTLEMENT AMOUNT"].eq(0),
                rincian_cmp["SELISIH"].eq(0),
            ],
            [
                "HANYA DI SETTLEMENT",
                "HANYA DI TIKET",
                "MATCH",
            ],
            default="SELISIH NOMINAL",
        )

        rincian_cmp = rincian_cmp[
            ["ORDER ID", "TARIF TIKET DETAIL", "SETTLEMENT AMOUNT", "SELISIH", "STATUS"]
        ]

        rincian_cmp = rincian_cmp[rincian_cmp["SELISIH"].ne(0)].copy()
        rincian_cmp["__abs__"] = rincian_cmp["SELISIH"].abs()
        rincian_cmp = rincian_cmp.sort_values(["__abs__", "ORDER ID"], ascending=[False, True], kind="stable")
        rincian_cmp = rincian_cmp.drop(columns="__abs__").reset_index(drop=True)
        rincian_cmp.insert(0, "NO", range(1, len(rincian_cmp) + 1))

        total_row_gap = pd.DataFrame([{
            "NO": "",
            "ORDER ID": "TOTAL",
            "TARIF TIKET DETAIL": rincian_cmp["TARIF TIKET DETAIL"].sum() if not rincian_cmp.empty else 0.0,
            "SETTLEMENT AMOUNT": rincian_cmp["SETTLEMENT AMOUNT"].sum() if not rincian_cmp.empty else 0.0,
            "SELISIH": rincian_cmp["SELISIH"].sum() if not rincian_cmp.empty else 0.0,
            "STATUS": "",
        }])

        rincian_view = pd.concat([rincian_cmp, total_row_gap], ignore_index=True)

        rincian_fmt = rincian_view.copy()
        for c in ["TARIF TIKET DETAIL", "SETTLEMENT AMOUNT", "SELISIH"]:
            rincian_fmt[c] = rincian_fmt[c].apply(_idr_fmt)

        bio_gap = io.BytesIO()
        with pd.ExcelWriter(bio_gap, engine="openpyxl") as xw_gap:
            rincian_view.to_excel(xw_gap, index=False, sheet_name="Rincian_Selisih")
            rincian_fmt.to_excel(xw_gap, index=False, sheet_name="Rincian_Selisih_View")

        rincian_selisih_table = rincian_fmt
        rincian_selisih_excel_bytes = bio_gap.getvalue()

    periode = f"{y}-{m:02d}"
    zona_waktu = ticket_tz_mode
    st.session_state["HASIL"]["rekon"] = {"periode": periode, "zona_waktu": zona_waktu, "table": fmt, "excel_bytes": bio.getvalue()}
    if df2_fmt_mi is not None:
        st.session_state["HASIL"]["detail_tiket"] = {"periode": periode, "zona_waktu": zona_waktu, "table": df2_fmt_mi}
    else:
        st.session_state["HASIL"].pop("detail_tiket", None)

    if detail_settle_table is not None and detail_settle_excel_bytes is not None:
        st.session_state["HASIL"]["detail_settlement"] = {
            "periode": periode,
            "zona_waktu": zona_waktu,
            "table": detail_settle_table,
            "excel_bytes": detail_settle_excel_bytes,
        }
    else:
        st.session_state["HASIL"].pop("detail_settlement", None)

    if rincian_selisih_table is not None and rincian_selisih_excel_bytes is not None:
        st.session_state["HASIL"]["rincian_selisih"] = {
            "periode": periode,
            "zona_waktu": zona_waktu,
            "table": rincian_selisih_table,
            "excel_bytes": rincian_selisih_excel_bytes,
        }
    else:
        st.session_state["HASIL"].pop("rincian_selisih", None)

    st.success("Proses selesai. Hasil tersimpan (klik download tidak perlu proses ulang).")


# =========================
# RENDER HASIL TERSIMPAN
# =========================

hasil = st.session_state.get("HASIL", {})

if "rekon" in hasil:
    st.subheader("Hasil Rekonsiliasi per Tanggal (mengikuti bulan parameter)")
    st.caption(f"Periode tersimpan: {hasil['rekon']['periode']} | Zona waktu Created: {hasil['rekon'].get('zona_waktu', 'WIB')}")
    st.dataframe(_style_right(hasil["rekon"]["table"]), use_container_width=True, hide_index=True)

    st.download_button(
        "Unduh Excel Rekonsiliasi",
        data=hasil["rekon"]["excel_bytes"],
        file_name=f"rekonsiliasi_{hasil['rekon']['periode']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_rekon",
    )

if "detail_tiket" in hasil:
    st.subheader("Detail Tiket per Tanggal — TYPE: GO SHOW & ONLINE × SUB-TIPE (J) [HANYA PAID]")
    st.caption(f"Periode tersimpan: {hasil['detail_tiket']['periode']} | Zona waktu Created: {hasil['detail_tiket'].get('zona_waktu', 'WIB')}")
    st.dataframe(_style_right(hasil["detail_tiket"]["table"]), use_container_width=True, hide_index=True)

if "detail_settlement" in hasil:
    st.subheader("DETAIL SETTLEMENT REPORT")
    st.caption(f"Periode tersimpan: {hasil['detail_settlement']['periode']} | Zona waktu Created: {hasil['detail_settlement'].get('zona_waktu', 'WIB')}")
    st.dataframe(_style_right(hasil["detail_settlement"]["table"]), use_container_width=True, hide_index=True)

    st.download_button(
        "Unduh Excel (Detail Settlement)",
        data=hasil["detail_settlement"]["excel_bytes"],
        file_name=f"detail_settlement_{hasil['detail_settlement']['periode']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_detail_settlement",
    )


if "rincian_selisih" in hasil:
    st.subheader("RINCIAN SELISIH ORDER ID — Tiket Detail vs Settlement Dana")
    st.caption(f"Periode tersimpan: {hasil['rincian_selisih']['periode']} | Zona waktu Created: {hasil['rincian_selisih'].get('zona_waktu', 'WIB')}")
    st.dataframe(_style_right(hasil["rincian_selisih"]["table"]), use_container_width=True, hide_index=True)

    st.download_button(
        "Unduh Excel (Rincian Selisih)",
        data=hasil["rincian_selisih"]["excel_bytes"],
        file_name=f"rincian_selisih_{hasil['rincian_selisih']['periode']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_rincian_selisih",
    )

if not hasil:
    st.info("Silakan upload file, pilih bulan-tahun, lalu klik **Proses**.")
